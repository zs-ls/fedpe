import copy
import time
import torch
import random
import os
import numpy as np
import yaml
import logging
import matplotlib.pyplot as plt
import argparse
from openpyxl import Workbook, load_workbook

from collections import defaultdict
from data.dataload import DataSet
from algorithms.FedPE.server import Server
from algorithms.FedPE.client import Client
from config.get_config import get_config


def setup_seed(seed):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)


def sync_lists_to_excel(workbook, list, file_name):
    """
    将多个列表同步到 Excel 中，每个列表占一列。
    :param lists: 以列表为元素的列表，例如 [[1, 2, 3], [4, 5, 6], ...]
    """
    # 找到当前表中最大的行数，确保新数据从正确位置写入
    sheet = workbook.active
    max_row = sheet.max_row
    # print(max_row)
    for col_idx, value in enumerate(list, start=1):  # 每个列表作为一列
        sheet.cell(row=max_row + 1, column=col_idx, value=value)
    workbook.save(file_name)


def main():
    start_time = time.time()

    parser = argparse.ArgumentParser()
    parser.add_argument('--config', type=str, default='./config/config.yaml', help='配置文件路径')
    command_args = parser.parse_args()

    # print("解析配置参数")
    args = get_config(command_args.config)
    pro_name = ("{}".format(args.seed) + "-" +
                args.model_type + "-" +
                args.dataset + "-" +
                args.data_distribution + "-" +
                args.algorithm + "-" +
                args.client_sample_method +
                "_sample"
                )
    log_file = "./log/" + pro_name + ".log"
    logging.basicConfig(
        filename=log_file,
        filemode="w",
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(formatter)
    logging.getLogger().addHandler(console_handler)

    result_file = f"./result/{pro_name}.xlsx"
    # directory = os.path.dirname(result_file)
    # if not os.path.exists(directory):
    #     os.makedirs(directory)  # 创建父目录
    #     print(f"目录不存在，已创建: {directory}")
    #     open(result_file, 'w').close()

    try:
        # 如果文件存在，则加载
        workbook = load_workbook(result_file)
        print(f"成功加载文件: {result_file}")
    except FileNotFoundError:
        # 如果文件不存在，则创建新的工作簿和工作表
        print(f"文件不存在，创建文件: {result_file}")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Data"  # 设置表名
        # 添加标题行
        sheet.append(["Index", "mAP", "CR", "clients_param_sum", "all_CR_per_round", "all_valid_param_per_round"])
        workbook.save(result_file)

    logging.info("设置随机种子")
    setup_seed(args.seed)

    logging.info("选择进行训练的设备")
    device = torch.device('cuda:{}'.format(
        args.gpu) if torch.cuda.is_available() and args.gpu != -1 else 'cpu')

    logging.info("为每个客户端分配数据")
    dataset = DataSet(args)
    logging.info(f"{dataset.dataset}")
    all_clients_train_set, all_client_test_set = dataset.allocate()

    logging.info("初始化服务器端")
    server = Server(args, device)

    logging.info(f"初始化全局模型,分类数量：{dataset.class_num}")
    server.model_init(dataset.class_num)
    # print("全局模型总参数：{}".format(server.model_parameters_num))

    logging.info(f"初始化客户端，客户端数量: {args.client_num}")
    clients = {i: Client(i, all_clients_train_set[i], all_client_test_set[i],
                         copy.deepcopy(server.clients_net[i]), device, args) for i in range(args.client_num)}
    logging.info("开始训练")
    round_cnt = 0
    all_mAP = []
    all_CR = []
    all_clients_param_num = []
    max_mAP = 0.0
    mean_CR = 0.0
    while True:
        logging.info("客户端采样")
        server.clients_sample()

        logging.info("将模型下发给选中的客户端，并开始本地训练")
        clients_param_list = []
        for i in server.clients_index_list:
            clients_param_list.append(clients[i].train(round_cnt, copy.deepcopy(server.clients_net[i])))

        logging.info("接收本地客户端发送的模型，并进行全局聚合")
        server.aggregation(clients_param_list)

        logging.info("评估本轮聚合后的模型")
        mAP = 0.0
        CR = 0.0
        clients_param_sum = 0
        all_CR_per_round = []
        all_valid_param_per_round = []
        for i in range(len(clients_param_list)):
            mAP += clients_param_list[i]["mAP"]
            # print(clients_param_list[i]["CR"])
            CR += clients_param_list[i]["CR"]
            all_CR_per_round.append(str(clients_param_list[i]["CR"]))
            clients_param_sum += clients_param_list[i]["valid_param_num"]
            all_valid_param_per_round.append(str(clients_param_list[i]["valid_param_num"]))
            # print(clients_param_list[i]["mAP"])
        mAP /= len(clients_param_list)
        CR /= len(clients_param_list)
        clients_param_sum //= len(clients_param_list)

        log_info = '    Round-{:>2d} | mAP:{:.2f}% | CR:{:.2f} | mVP:{:d}'.format(
            round_cnt + 1, mAP * 100, CR, clients_param_sum
        )
        logging.info(log_info)

        # if (float(test_auc) - args.object_acc) > 0:
        #     exit()
        all_mAP.append(mAP)
        all_CR.append(CR)
        all_clients_param_num.append(clients_param_sum)
        # print("将数据保存到excel表中")
        sync_lists_to_excel(workbook, [round_cnt + 1, mAP, CR, clients_param_sum,
                             ",".join(all_CR_per_round), ",".join(all_valid_param_per_round)],
                            result_file)
        round_cnt += 1
        if round_cnt >= args.total_round:
            end_time = time.time()
            training_time = end_time - start_time
            # 将训练时间转换为小时、分钟、秒
            hours, remainder = divmod(training_time, 3600)
            minutes, seconds = divmod(remainder, 60)

            # 使用logging输出训练时间
            logging.info(f"Training Time: {int(hours)} hours {int(minutes)} minutes {int(seconds)} seconds")

            combined = list(zip(all_mAP, all_CR))
            max_mAP = max(all_mAP)
            max_mAP_ind = all_mAP.index(max_mAP)
            max_CR = max(all_CR)
            max_CR_ind = all_CR.index(max_CR)

            info = 'The best mAP: {:.2f}% | CR: {:.2f} | param_num:{} | complete_param_num:{} | Round: {:>2d}'.format(
                combined[max_mAP_ind][0] * 100, combined[max_mAP_ind][1], all_clients_param_num[max_mAP_ind],
                server.model_parameters_num, max_mAP_ind + 1
            )
            logging.info(info)
            info = 'The best CR: {:.2f} | mAP: {:.2f}% | param_num:{} | complete_param_num:{} | Round: {:>2d}'.format(
                combined[max_CR_ind][1], combined[max_CR_ind][0] * 100, all_clients_param_num[max_mAP_ind],
                server.model_parameters_num, max_CR_ind + 1
            )
            logging.info(info)
            # 创建子图，设置布局为 1 行 2 列
            fig, axs = plt.subplots(1, 2, figsize=(12, 5))  # figsize 控制整体窗口大小

            axs[0].plot(range(1, round_cnt + 1), all_mAP, label=args.algorithm, color='red')
            axs[0].set_title(log_file)
            axs[0].set_xlabel('Communication Rounds')
            axs[0].set_ylabel('Test mAP')
            # plt.title('Test Accuracy vs Round')
            axs[0].legend()
            axs[0].grid(True)
            axs[1].plot(range(1, round_cnt + 1), all_clients_param_num, label='FedPE', color='red')
            axs[1].set_xlabel('Communication Rounds')
            axs[1].set_ylabel('Number of transmitted parameters of subnets')
            axs[1].set_title(log_file)
            axs[1].legend()
            axs[1].grid(True)
            # 调整子图布局
            plt.tight_layout()  # 自动调整布局以防止重叠
            plt.show()
            exit()


if __name__ == "__main__":
    main()


